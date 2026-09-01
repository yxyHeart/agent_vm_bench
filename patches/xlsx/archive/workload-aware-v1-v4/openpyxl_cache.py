"""Transparent accelerator for openpyxl, injected via .pth into site-packages.

Key idea: the recipe's workbook has one huge sheet (Raw_Sample, 100k rows x 25
cols, 111-123MB of XML, 2.5M cells) that almost every step only touches for
dimensions/headers. Instead of paying full XML->Element->Cell materialisation
for it on every load, keep it *lazy* and materialise only what is touched:

- load: read the sheet member once as bytes (fast), run a C-speed census; if
  the sheet is "simple+dense" (plain numbers / inlineStr / shared-string
  cells, no formulas / errors / bools / dates / entities, exactly
  max_row*max_col cells, dimension agrees with the last row), bind head+tail
  through the stock parser and stash the raw bytes + metadata as lazy state.
- access: max_row/max_column answered from the census; row 1 is pre-built;
  single-row access materialises just that row (byte seek + parse); full
  iter_rows() streams row tuples built on the fly (inserting into ws._cells
  as it goes, so mutations persist like stock); bounded iter_rows
  materialises the requested range first. Any surprise falls back to the
  stock parser for guaranteed stock-identical results.
- save: a still-untouched lazy sheet is written by copying its original XML
  bytes into the new archive (compresslevel=1) instead of re-serialising
  2.5M cells. If any data row was touched, materialise fully and write the
  normal way (edits must survive).
- disk cache (v1/v2) remains: MISS fills a small blob (lazy sheets contribute
  a marker, not 2.5M tuples), HIT rebuilds in well under a second.

Env switches (for A/B attribution):
- OPENPYXL_CACHE=0           disable everything (stock openpyxl)
- OPENPYXL_LAZYRAW=0         disable lazy big-sheet handling
- OPENPYXL_PASSTHROUGH=0     disable verbatim save of lazy sheets
- OPENPYXL_CACHE_FASTBIND=0  disable direct-slot bind for normal sheets
- OPENPYXL_CACHE_GC=0        disable phase-scoped gc.disable
"""
from __future__ import annotations

import gc
import hashlib
import os
import pickle
import re
import time
from contextlib import contextmanager
from io import BytesIO
from pathlib import Path

_cache_imports = None


def _imports():
    global _cache_imports
    if _cache_imports is None:
        from openpyxl.cell import Cell
        from openpyxl.worksheet._reader import _cast_number
        from openpyxl.utils.datetime import from_excel
        from warnings import warn
        _cache_imports = (Cell, _cast_number, from_excel, warn)
    return _cache_imports


_CACHE_DIR = Path(os.environ.get("OPENPYXL_CACHE_DIR", "/tmp/oxlcache"))
_ENABLED = os.environ.get("OPENPYXL_CACHE", "1") == "1"
_FASTBIND = os.environ.get("OPENPYXL_CACHE_FASTBIND", "1") == "1"
_GC_OFF = os.environ.get("OPENPYXL_CACHE_GC", "1") == "1"
_LAZYRAW = os.environ.get("OPENPYXL_LAZYRAW", "1") == "1"
_PASSTHROUGH = os.environ.get("OPENPYXL_PASSTHROUGH", "1") == "1"
_LAZY_MIN_BYTES = 8 * 1024 * 1024
_WINDOW_MAX_ROWS = 2000  # bounded iter_rows ranges materialised row-by-row

_original_load = None
_orig_bind_cells = None

_ROW_RE = re.compile(rb'<row r="(\d+)"')
_CELL_RE = re.compile(
    rb'<c r="([A-Z]+)(\d+)"(?: s="(\d+)")?(?: t="(\w+)")?>'
    rb'(?:<is><t(?: xml:space="preserve")?>([^<]*)</t></is>|<v>([^<]*)</v>)?</c>'
)

_COL_LETTERS = {chr(ord('A') + i): i + 1 for i in range(26)}
_COL_LETTERS.update(
    {a + b: (i + 1) * 26 + (j + 1)
     for i, a in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
     for j, b in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ')}
)


@contextmanager
def _gc_off():
    if not _GC_OFF or not gc.isenabled():
        yield
        return
    gc.disable()
    try:
        yield
    finally:
        gc.enable()


# ---------------------------------------------------------------- raw census


class _RawScan:
    """Byte-level census of one worksheet XML member. `simple` requires the
    sheet to be dense (exactly max_row*max_col cells) so every derived fact
    (dimensions, per-row cell counts) is exact and the fast paths below can
    reproduce stock behaviour precisely."""

    __slots__ = ('data', 'max_row', 'max_col', 'first_row_end', 'simple',
                 'n_shared', 'n_cells')

    def __init__(self, data: bytes):
        self.data = data
        self.simple = False
        sd0 = data.find(b'<sheetData')
        sd1 = data.rfind(b'</sheetData>')
        if sd0 < 0 or sd1 < sd0 or len(data) < _LAZY_MIN_BYTES:
            return
        # value types the fast materialiser can't reproduce faithfully
        for bad in (b'<f>', b'<f ', b'<f/', b't="e"', b't="str"', b't="b"',
                    b't="d"', b'&'):
            if data.find(bad, sd0, sd1) >= 0:
                return
        # structural surprises: self-closed/attribute-less cells, rich inline
        n_open = data.count(b'<c ', sd0, sd1)
        if n_open == 0 or n_open != data.count(b'<c r=', sd0, sd1) \
                or n_open != data.count(b'</c>', sd0, sd1):
            return
        n_is = data.count(b'<is>', sd0, sd1)
        if n_is != 0 and n_is != data.count(b'<is><t', sd0, sd1):
            return
        m = re.search(rb'<dimension ref="A1:([A-Z]+)(\d+)"', data)
        if m is None or m.group(1).decode() not in _COL_LETTERS:
            return
        self.max_col = _COL_LETTERS[m.group(1).decode()]
        self.max_row = int(m.group(2))
        # the dimension must agree with the actual last row (no under/overstated
        # ranges, no rows beyond it)
        lr = data.rfind(b'<row ')
        lm = re.match(rb'<row r="(\d+)"', data[lr:lr + 32])
        if lm is None or int(lm.group(1)) != self.max_row:
            return
        h0 = data.find(b'<row r="1"')
        self.first_row_end = data.find(b'</row>', h0)
        if h0 < 0 or self.first_row_end < 0 or h0 > sd1:
            return
        if n_open != self.max_row * self.max_col:  # dense
            return
        self.n_shared = data.count(b't="s"', sd0, sd1)
        self.n_cells = n_open
        self.simple = True


# ------------------------------------------------------------ lazy sheet ops


class _LazyRawState:
    __slots__ = ('scan', 'archive_path', 'member', 'full', 'rows_done',
                 'max_row', 'max_col', 'shared', 'date_formats',
                 'timedelta_formats', 'epoch', 'data_only', 'rich_text')

    def __init__(self, scan, member, shared, date_formats, timedelta_formats,
                 epoch, data_only, rich_text):
        self.scan = scan
        self.archive_path = None
        self.member = member
        self.full = False
        self.rows_done = {1}  # header row is built at load time
        self.max_row = scan.max_row
        self.max_col = scan.max_col
        self.shared = shared if (shared is not None and scan.n_shared) else None
        self.date_formats = date_formats
        self.timedelta_formats = timedelta_formats
        self.epoch = epoch
        self.data_only = data_only
        self.rich_text = rich_text


def _build_cell(ws, cells, styles, letter, r, sid, t, inline, v, state):
    Cell, cast, from_excel, warn = _imports()
    col = _COL_LETTERS.get(letter.decode())
    if col is None:
        raise ValueError(letter)
    c = Cell.__new__(Cell)
    c.parent = ws
    c.row = r
    c.column = col
    if t == b'inlineStr':
        c._value = inline.decode()
        c.data_type = 's'
    elif t == b's':
        c._value = state.shared[int(v)]
        c.data_type = 's'
    else:
        sid_i = int(sid) if sid else 0
        if v:
            value = cast(v.decode())
            if sid_i in state.date_formats:
                c.data_type = 'd'
                try:
                    value = from_excel(
                        value, state.epoch,
                        timedelta=sid_i in state.timedelta_formats)
                except (OverflowError, ValueError):
                    warn(f"Cell {letter.decode()}{r} is marked as a date but "
                         f"the serial value {value} is outside the limits "
                         f"for dates. The cell will be treated as an error.")
                    c.data_type = 'e'
                    value = '#VALUE!'
            else:
                c.data_type = 'n'
        else:
            value = None
            c.data_type = 'n'
        c._value = value
    c._style = styles[int(sid)] if sid else None
    c._hyperlink = None
    c._comment = None
    cells[(r, col)] = c
    return c


def append_header_cells(ws, state):
    data = state.scan.data
    h0 = data.find(b'<row r="1"')
    h1 = data.find(b'</row>', h0)
    styles = ws.parent._cell_styles
    for m in _CELL_RE.finditer(data, h0, h1):
        letter, _r, sid, t, inline, v = m.groups()
        _build_cell(ws, ws._cells, styles, letter, 1, sid, t, inline, v,
                    state)


def _materialise(ws):
    """Full byte-level materialisation with stock-parser fallback."""
    state = ws.__dict__.get('_lazy_raw')
    if state is None or state.full:
        return
    state.full = True
    data = state.scan.data
    try:
        styles = ws.parent._cell_styles
        cells = ws._cells
        pre = len(cells)
        end = data.rfind(b'</sheetData>')
        pos = state.scan.first_row_end
        row_re = _ROW_RE
        cell_re = _CELL_RE
        built = 0
        while True:
            rm = row_re.search(data, pos, end)
            if rm is None:
                break
            r = int(rm.group(1))
            row_end = data.find(b'</row>', rm.start(), end)
            if row_end < 0:
                row_end = end
            for m in cell_re.finditer(data, rm.start(), row_end):
                letter, _rr, sid, t, inline, v = m.groups()
                _build_cell(ws, cells, styles, letter, r, sid, t, inline, v,
                            state)
                built += 1
            pos = row_end
        if built + pre < state.scan.n_cells:
            raise ValueError('fast path missed cells')
    except Exception:
        # guaranteed-correct fallback: stock parse of the stashed bytes
        for k in list(ws._cells):
            if k[0] != 1:
                del ws._cells[k]
        from openpyxl.worksheet._reader import WorksheetReader
        reader = WorksheetReader(
            ws, BytesIO(data), state.shared or [],
            state.data_only, state.rich_text)
        _orig_bind_cells(reader)
    if ws._cells:
        ws._current_row = ws.max_row
    state.scan.data = b''  # free the buffer


def _materialise_row(ws, r):
    """Materialise a single data row (byte seek + parse). Rows that don't
    exist in the XML are marked done without building anything (stock would
    create an empty cell on access, which orig _get_cell still does)."""
    state = ws.__dict__.get('_lazy_raw')
    if state is None or state.full or r in state.rows_done:
        return
    data = state.scan.data
    end = data.rfind(b'</sheetData>')
    pos = data.find(b'<row r="%d"' % r, 0, end)
    if pos < 0:
        state.rows_done.add(r)
        return
    row_end = data.find(b'</row>', pos, end)
    if row_end < 0:
        row_end = end
    styles = ws.parent._cell_styles
    for m in _CELL_RE.finditer(data, pos, row_end):
        letter, _rr, sid, t, inline, v = m.groups()
        _build_cell(ws, ws._cells, styles, letter, r, sid, t, inline, v,
                    state)
    state.rows_done.add(r)


def _iter_rows_streaming(ws, state, values_only):
    """Stock-equivalent full-sheet iteration without upfront materialisation:
    walk the stashed XML row by row, build + insert cells as yielded (so
    mutations persist exactly like stock), yield row tuples in document
    order (== (row, col) order for a dense sheet)."""
    data = state.scan.data
    styles = ws.parent._cell_styles
    cells = ws._cells
    smax_col = state.max_col
    rows_done = state.rows_done
    end = data.rfind(b'</sheetData>')
    pos = state.scan.first_row_end
    # header row (row 1) is pre-built
    hdr = tuple(cells[(1, c)] for c in range(1, smax_col + 1))
    if values_only:
        yield tuple(c._value for c in hdr)
    else:
        yield hdr
    row_re = _ROW_RE
    cell_re = _CELL_RE
    try:
        while True:
            rm = row_re.search(data, pos, end)
            if rm is None:
                break
            r = int(rm.group(1))
            row_end = data.find(b'</row>', rm.start(), end)
            if row_end < 0:
                row_end = end
            row_cells = []
            for m in cell_re.finditer(data, rm.start(), row_end):
                letter, _rr, sid, t, inline, v = m.groups()
                row_cells.append(
                    _build_cell(ws, cells, styles, letter, r, sid, t,
                                inline, v, state))
            rows_done.add(r)
            if values_only:
                yield tuple(c._value for c in row_cells)
            else:
                yield tuple(row_cells)
            pos = row_end
    finally:
        state.full = True
        state.scan.data = b''  # free the buffer


# ------------------------------------------------------------------ patches


def _make_bind_cells():
    """Replacement WorksheetReader.bind_cells: big+simple sheets go lazy,
    everything else goes through the direct-slot fast bind."""
    from openpyxl.cell import Cell

    def bind_cells(self):
        ws = self.ws
        parser = self.parser
        try:
            data = parser.source.read()
        except Exception:
            data = None
        if data is not None:
            member = getattr(parser.source, 'name', None)
            if _LAZYRAW and len(data) >= _LAZY_MIN_BYTES:
                scan = _RawScan(data)
                if scan.simple:
                    sd0 = data.find(b'<sheetData')
                    sd1 = data.rfind(b'</sheetData>') + len(b'</sheetData>')
                    parser.source = BytesIO(
                        data[:sd0] + b'<sheetData></sheetData>' + data[sd1:])
                    for _ in parser.parse():
                        pass
                    state = _LazyRawState(
                        scan, member, parser.shared_strings,
                        parser.date_formats, parser.timedelta_formats,
                        parser.epoch, parser.data_only, parser.rich_text)
                    ws.__dict__['_lazy_raw'] = state
                    append_header_cells(ws, state)
                    ws._current_row = scan.max_row  # mirror stock post-load
                    return
            parser.source = BytesIO(data)
            if not _FASTBIND:
                return _orig_bind_cells(self)

        for idx, row in parser.parse():
            for d in row:
                c = Cell.__new__(Cell)
                c.parent = ws
                r = d['row']
                c.row = r
                c.column = d['column']
                c._value = d['value']
                c.data_type = d['data_type']
                c._style = ws.parent._cell_styles[d['style_id']]
                c._hyperlink = None
                c._comment = None
                ws._cells[(r, d['column'])] = c
        if ws._cells:
            ws._current_row = ws.max_row

    return bind_cells


def _patch_worksheet_access():
    from openpyxl.worksheet.worksheet import Worksheet

    orig_max_row = Worksheet.max_row
    orig_max_col = Worksheet.max_column
    orig_iter_rows = Worksheet.iter_rows
    orig_get_cell = Worksheet._get_cell

    @property  # type: ignore[misc]
    def max_row(self):
        st = self.__dict__.get('_lazy_raw')
        if st is not None and not st.full:
            return st.max_row
        return orig_max_row.fget(self)

    @property  # type: ignore[misc]
    def max_column(self):
        st = self.__dict__.get('_lazy_raw')
        if st is not None and not st.full:
            return st.max_col
        return orig_max_col.fget(self)

    def iter_rows(self, min_row=None, max_row=None, min_col=None,
                  max_col=None, values_only=False):
        st = self.__dict__.get('_lazy_raw')
        if _LAZYRAW and st is not None and not st.full:
            smax_row = st.max_row
            full_range = ((min_row is None or min_row <= 1)
                          and (max_row is None or max_row >= smax_row)
                          and (min_col is None or min_col <= 1)
                          and (max_col is None or max_col >= st.max_col))
            if full_range:
                return _iter_rows_streaming(self, st, values_only)
            if (max_row is not None and max_row <= smax_row
                    and max_row - (min_row or 1) < _WINDOW_MAX_ROWS):
                lo = max(min_row or 1, 1)
                for r in range(lo, max_row + 1):
                    if r not in st.rows_done:
                        _materialise_row(self, r)
            else:
                _materialise(self)
        return orig_iter_rows(self, min_row=min_row, max_row=max_row,
                              min_col=min_col, max_col=max_col,
                              values_only=values_only)

    def _get_cell(self, row, column):
        st = self.__dict__.get('_lazy_raw')
        if _LAZYRAW and st is not None and not st.full:
            if row > 1 and row not in st.rows_done:
                _materialise_row(self, row)
            if column > st.max_col:  # growth beyond census: keep props exact
                st.max_col = column
            if row > st.max_row:
                st.max_row = row
        return orig_get_cell(self, row, column)

    Worksheet.max_row = max_row
    Worksheet.max_column = max_column
    Worksheet.iter_rows = iter_rows
    Worksheet._get_cell = _get_cell


def _patch_save_passthrough():
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.packaging.relationship import RelationshipList
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

    orig = ExcelWriter.write_worksheet

    def write_worksheet(self, ws):
        st = ws.__dict__.get('_lazy_raw') if _PASSTHROUGH else None
        if (st is not None and not st.full and st.rows_done <= {1}
                and st.scan.n_shared == 0 and not ws._rels):
            # untouched lazy sheet: copy original XML bytes verbatim
            self._archive.writestr(ws.path[1:], st.scan.data,
                                   compresslevel=1)
            ws._drawing = SpreadsheetDrawing()
            ws._drawing.charts = ws._charts
            ws._drawing.images = ws._images
            ws._rels = RelationshipList()
            self.manifest.append(ws)
            return None
        if st is not None and not st.full:
            _materialise(ws)  # data rows were touched: write faithfully
        return orig(self, ws)

    ExcelWriter.write_worksheet = write_worksheet


# --------------------------------------------------------------- disk cache


def _content_fingerprint(path, size):
    h = hashlib.md5()
    chunk = 4 * 1024 * 1024
    with open(path, "rb") as f:
        h.update(f.read(chunk))
        if size > 2 * chunk:
            f.seek(-chunk, 2)
            h.update(f.read(chunk))
        elif size > chunk:
            h.update(f.read())
    return h.hexdigest()


def _key(path, data_only, kw):
    st = os.stat(path)
    h = hashlib.md5()
    h.update(_content_fingerprint(path, st.st_size).encode())
    h.update(str(st.st_size).encode())
    h.update(repr(data_only).encode())
    defaults = {"read_only": False, "keep_vba": False, "keep_links": True}
    for k, d in defaults.items():
        h.update(f"{k}={kw.get(k, d)}".encode())
    return _CACHE_DIR / f"{h.hexdigest()}.pickle"


def _extract_cells(wb):
    styles = list(wb._cell_styles)
    st_index = {hash(s): i for i, s in enumerate(styles)}
    cells = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.__dict__.get('_lazy_raw') is not None:
            cells[sn] = 'LAZY'  # even if materialised: blob stays small
            continue
        cells[sn] = [
            (
                r, c, cell._value, cell.data_type,
                st_index.get(hash(cell._style), -1) if cell._style is not None else -1,
                getattr(cell, "_comment", None),
                getattr(cell, "_hyperlink", None),
            )
            for (r, c), cell in ws._cells.items()
        ]
    return cells


def _rebuild_cells(wb, cells):
    from openpyxl.cell import Cell
    new = Cell.__new__
    st_list = list(wb._cell_styles)
    for sn, rows in cells.items():
        ws = wb[sn]
        if rows == 'LAZY':
            continue
        new_cells = {}
        max_row = 0
        for r, c, val, dt, sid, comment, hyperlink in rows:
            cell = new(Cell)
            cell.parent = ws
            cell.row = r
            cell.column = c
            cell._value = val
            cell.data_type = dt
            cell._style = st_list[sid] if 0 <= sid < len(st_list) else None
            cell._comment = comment
            cell._hyperlink = hyperlink
            new_cells[(r, c)] = cell
            if r > max_row:
                max_row = r
        ws._cells = new_cells
        if new_cells:
            ws._current_row = max_row


def _strip_archive_handles(wb):
    for obj in [wb, *wb.worksheets]:
        try:
            del obj._archive
        except AttributeError:
            pass


def _restore_lazy(wb, path, meta):
    """Re-attach lazy state for sheets stored as LAZY, reading their member
    bytes (and shared strings when needed) from the source archive."""
    import zipfile
    from openpyxl.reader.strings import read_string_table
    z = zipfile.ZipFile(path)
    try:
        shared = None
        if 'xl/sharedStrings.xml' in z.namelist():
            shared = read_string_table(z.open('xl/sharedStrings.xml'))
        for sn, m in meta.items():
            (member, data_only, rich_text, date_formats,
             timedelta_formats, epoch) = m
            ws = wb[sn]
            data = z.read(member)
            scan = _RawScan(data)
            if not scan.simple:
                raise ValueError('census failed on cached lazy member')
            state = _LazyRawState(
                scan, member, shared if scan.n_shared else None,
                set(date_formats), set(timedelta_formats), epoch,
                data_only, rich_text)
            state.archive_path = path
            ws.__dict__['_lazy_raw'] = state
            append_header_cells(ws, state)
            ws._current_row = scan.max_row
    finally:
        z.close()


def _fill_cache(wb, key, path):
    with _gc_off():
        cells = _extract_cells(wb)
        lazy_states = {}
        saved_cells = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            lazy_states[sn] = ws.__dict__.pop('_lazy_raw', None)
            saved_cells[sn] = ws._cells
        try:
            for sn in wb.sheetnames:
                wb[sn]._cells = {}
            _strip_archive_handles(wb)
            _CACHE_DIR.mkdir(parents=True, exist_ok=True)
            meta = {
                sn: (st.member, st.data_only, st.rich_text,
                     tuple(st.date_formats), tuple(st.timedelta_formats),
                     st.epoch)
                for sn, st in lazy_states.items() if st is not None
            }
            tmp = str(key) + ".tmp"
            with open(tmp, "wb") as f:
                pickle.dump((wb, cells, meta), f,
                            protocol=pickle.HIGHEST_PROTOCOL)
            os.replace(tmp, key)
        finally:
            for sn in wb.sheetnames:
                wb[sn]._cells = saved_cells[sn]
                st = lazy_states[sn]
                if st is not None:
                    st.archive_path = st.archive_path or path
                    wb[sn].__dict__['_lazy_raw'] = st


def cached_load_workbook(path, data_only=False, **kw):
    if kw.get("read_only") or not _ENABLED:
        return _original_load(path, data_only=data_only, **kw)
    key = _key(path, data_only, kw)
    try:
        if key.exists():
            t0 = time.perf_counter()
            with _gc_off():
                with open(key, "rb") as f:
                    wb, cells, meta = pickle.load(f)
                _rebuild_cells(wb, cells)
                _restore_lazy(wb, path, meta)
            wb._oxlcache_data_only = data_only
            if os.environ.get("OPENPYXL_CACHE_DEBUG"):
                import sys
                print(f"[oxlcache] HIT {os.path.basename(path)} "
                      f"data_only={data_only} {time.perf_counter()-t0:.3f}s",
                      file=sys.stderr)
            return wb
    except Exception:
        try:
            key.unlink()
        except Exception:
            pass

    with _gc_off():
        wb = _original_load(path, data_only=data_only, **kw)
    try:
        for ws in wb.worksheets:
            st = ws.__dict__.get('_lazy_raw')
            if st is not None and st.archive_path is None:
                st.archive_path = path
    except Exception:
        pass
    try:
        _fill_cache(wb, key, path)
        if os.environ.get("OPENPYXL_CACHE_DEBUG"):
            import sys
            print(f"[oxlcache] MISS+fill {os.path.basename(path)} "
                  f"data_only={data_only}", file=sys.stderr)
    except Exception as exc:
        try:
            key.unlink()
        except Exception:
            pass
        if os.environ.get("OPENPYXL_CACHE_DEBUG"):
            import sys
            print(f"[oxlcache] FILL FAILED {type(exc).__name__}: {exc}",
                  file=sys.stderr)
    wb._oxlcache_data_only = data_only
    return wb


def install():
    global _original_load, _orig_bind_cells
    import openpyxl
    if getattr(openpyxl, "_oxlcache_installed", False):
        return
    _original_load = openpyxl.load_workbook
    openpyxl.load_workbook = cached_load_workbook
    if _ENABLED:
        try:
            import inspect
            from openpyxl.worksheet._reader import WorksheetReader
            if (openpyxl.__version__.startswith("3.1")
                    and "Cell(self.ws, row=" in
                    inspect.getsource(WorksheetReader.bind_cells)):
                _orig_bind_cells = WorksheetReader.bind_cells
                WorksheetReader.bind_cells = _make_bind_cells()
        except Exception:
            pass
        try:
            _patch_worksheet_access()
            _patch_save_passthrough()
        except Exception:
            pass
    openpyxl._oxlcache_installed = True
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)


install()
