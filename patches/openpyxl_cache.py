"""On-disk cache + interpreter-level fast paths for openpyxl (read_only=False).

Transparent layer injected via oxlcache.pth (site-packages): every `python3`
auto-patches openpyxl.load_workbook. Recipe commands unchanged.

Levers (each independently kill-switchable via env for A/B attribution):
- disk cache (v1): MISS fills a pickled cell-less wb + cell table; HIT skips
  XML parse and rebuilds the 2.5M Cell objects from the table.
- GC off (v2): during load/rebuild/extract the heap grows monotonically toward
  2.5M live objects; periodic cyclic-GC scans over them are pure CPU waste.
  Env OPENPYXL_CACHE_GC=0 disables.
- direct-slot Cell construction (v2): Cell.__new__ + slot assignment instead
  of Cell.__init__, both on cache HIT rebuild and on openpyxl's own parse path
  (WorksheetReader.bind_cells); StyleArray table entries are shared instead of
  copied per cell (StyleArray is treated as immutable by openpyxl).
  Env OPENPYXL_CACHE_FASTBIND=0 disables the parse-path patch.
- lxml writer (v2, image-level): openpyxl 3.1.5 flips write_cell/xmlfile to
  lxml at import when lxml is installed (module-level dispatch), while the
  read path (iterparse) is unconditionally stdlib ElementTree — so installing
  lxml speeds up saves without touching loads. Env OPENPYXL_LXML=False (stock
  openpyxl flag) reverts the writer to etree.
- v2 dropped the save-path cache fill: in recipe v2 no openpyxl load ever
  reads a freshly saved file (the only openpyxl save, TP-07's enhance output,
  is immediately rewritten by LibreOffice recalc), so filling on save was
  wasted CPU on the critical path.
- optional gc.freeze() after a MISS load (env OPENPYXL_CACHE_FREEZE=1, default
  off): excludes the loaded object tree from all later GC scans inside the
  same process (helps long-lived multi-load scripts).

Result (2-core/4G container, fixed single task): v1 259s -> 174.9s (-33%);
v2 see docs/xlsx-cache-report.md update.
"""
from __future__ import annotations

import gc
import hashlib
import inspect
import os
import pickle
import time
from contextlib import contextmanager
from pathlib import Path

_CACHE_DIR = Path(os.environ.get("OPENPYXL_CACHE_DIR", "/tmp/oxlcache"))
_ENABLED = os.environ.get("OPENPYXL_CACHE", "1") == "1"
_FASTBIND = os.environ.get("OPENPYXL_CACHE_FASTBIND", "1") == "1"
_GC_OFF = os.environ.get("OPENPYXL_CACHE_GC", "1") == "1"
_FREEZE = os.environ.get("OPENPYXL_CACHE_FREEZE", "0") == "1"
_original_load = None


@contextmanager
def _gc_off():
    """Disable cyclic GC for the wrapped phase; restore the prior state after.

    The phases wrapped here (parse / rebuild / extract) allocate millions of
    long-lived objects and almost no cyclic garbage: gen0/1 threshold trips
    repeatedly rescan the growing live set for nothing."""
    if not _GC_OFF or not gc.isenabled():
        yield
        return
    gc.disable()
    try:
        yield
    finally:
        gc.enable()


def _content_fingerprint(path, size):
    """md5 of first+last 4MB. ~20ms for 123MB. Cross-path share of identical
    content; invalidates on real content change. No mtime, no same-second hits."""
    h = hashlib.md5()
    chunk = 4 * 1024 * 1024
    with open(path, "rb") as f:
        h.update(f.read(chunk))
        if size > 2 * chunk:
            f.seek(-chunk, 2)
            h.update(f.read(chunk))
        elif size > chunk:
            h.update(f.read())
    return h.hexdigest()


def _key(path, data_only, kw):
    st = os.stat(path)
    h = hashlib.md5()
    h.update(_content_fingerprint(path, st.st_size).encode())
    h.update(str(st.st_size).encode())
    h.update(repr(data_only).encode())
    # normalize kwargs to real defaults so "omitted" (default) and "explicitly
    # passed default" hash the same (read_only=None vs read_only=False are the
    # same non-read-only load).
    defaults = {"read_only": False, "keep_vba": False, "keep_links": True}
    for k, d in defaults.items():
        h.update(f"{k}={kw.get(k, d)}".encode())
    return _CACHE_DIR / f"{h.hexdigest()}.pickle"


def _extract_cells(wb):
    """{sheetname: [(row, col, _value, data_type, style_id, comment, hyperlink), ...]}.
    style_id is the value-hash index of cell._style in wb._cell_styles.
    comment/hyperlink are stored so checks reading cell.comment / cell.hyperlink
    still pass on a hit (the cell-less wb pickle alone drops these per-cell refs)."""
    styles = list(wb._cell_styles)
    st_index = {hash(s): i for i, s in enumerate(styles)}
    cells = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cells[sn] = [
            (
                r, c, cell._value, cell.data_type,
                st_index.get(hash(cell._style), -1) if cell._style is not None else -1,
                getattr(cell, "_comment", None),
                getattr(cell, "_hyperlink", None),  # MergedCell has no _hyperlink slot
            )
            for (r, c), cell in ws._cells.items()
        ]
    return cells


def _rebuild_cells(wb, cells):
    """Direct-slot rebuild: Cell.__new__ + slot assignment (skips Cell.__init__
    re-initialisation and the per-cell StyleArray copy)."""
    from openpyxl.cell import Cell
    new = Cell.__new__
    st_list = list(wb._cell_styles)
    for sn, rows in cells.items():
        ws = wb[sn]
        new_cells = {}
        max_row = 0
        for r, c, val, dt, sid, comment, hyperlink in rows:
            cell = new(Cell)
            cell.parent = ws
            cell.row = r
            cell.column = c
            cell._value = val
            cell.data_type = dt
            cell._style = st_list[sid] if 0 <= sid < len(st_list) else None
            cell._comment = comment
            cell._hyperlink = hyperlink
            new_cells[(r, c)] = cell
            if r > max_row:
                max_row = r
        ws._cells = new_cells
        if new_cells:
            ws._current_row = max_row


def _strip_archive_handles(wb):
    """Delete read_only zip handles so the cell-less wb pickles cleanly. In
    normal mode _archive is absent (close() guards via hasattr), so this is a
    no-op; never *create* the attribute (a spurious None breaks close())."""
    for obj in [wb, *wb.worksheets]:
        try:
            del obj._archive
        except AttributeError:
            pass


def _fill_cache(wb, key):
    """Extract cells, pickle cell-less wb + cells (atomic tmp+rename), restore wb._cells."""
    with _gc_off():
        cells = _extract_cells(wb)
        saved_cells = {sn: wb[sn]._cells for sn in wb.sheetnames}
        try:
            for sn in wb.sheetnames:
                wb[sn]._cells = {}
            _strip_archive_handles(wb)
            _CACHE_DIR.mkdir(parents=True, exist_ok=True)
            tmp = str(key) + ".tmp"
            with open(tmp, "wb") as f:
                pickle.dump((wb, cells), f, protocol=pickle.HIGHEST_PROTOCOL)
            os.replace(tmp, key)
        finally:
            for sn in wb.sheetnames:
                wb[sn]._cells = saved_cells[sn]


def _make_fast_bind_cells():
    """Replacement for WorksheetReader.bind_cells (openpyxl 3.1.x): identical
    semantics — same slots end-state, same ws._cells keys, same _current_row —
    but builds Cells via __new__ + direct slot writes and shares the StyleArray
    table entry instead of copying it per cell."""
    from openpyxl.cell import Cell

    def bind_cells(self):
        ws = self.ws
        styles = ws.parent._cell_styles
        cells = ws._cells
        new = Cell.__new__
        for idx, row in self.parser.parse():
            for d in row:
                c = new(Cell)
                c.parent = ws
                r = d['row']
                c.row = r
                c.column = d['column']
                c._value = d['value']
                c.data_type = d['data_type']
                c._style = styles[d['style_id']]
                c._hyperlink = None
                c._comment = None
                cells[(r, d['column'])] = c
        if cells:
            ws._current_row = ws.max_row

    return bind_cells


def cached_load_workbook(path, data_only=False, **kw):
    if kw.get("read_only") or not _ENABLED:
        return _original_load(path, data_only=data_only, **kw)
    key = _key(path, data_only, kw)
    try:
        if key.exists():
            t0 = time.perf_counter()
            with _gc_off():
                with open(key, "rb") as f:
                    wb, cells = pickle.load(f)
                _rebuild_cells(wb, cells)
            wb._oxlcache_data_only = data_only
            if os.environ.get("OPENPYXL_CACHE_DEBUG"):
                import sys
                print(f"[oxlcache] HIT {os.path.basename(path)} data_only={data_only} "
                      f"{time.perf_counter()-t0:.3f}s", file=sys.stderr)
            return wb
    except Exception:
        try:
            key.unlink()
        except Exception:
            pass

    with _gc_off():
        wb = _original_load(path, data_only=data_only, **kw)
    if _FREEZE:
        try:
            gc.freeze()
        except Exception:
            pass
    try:
        _fill_cache(wb, key)
        if os.environ.get("OPENPYXL_CACHE_DEBUG"):
            import sys
            print(f"[oxlcache] MISS+fill {os.path.basename(path)} data_only={data_only}",
                  file=sys.stderr)
    except Exception as exc:
        try:
            key.unlink()
        except Exception:
            pass
        if os.environ.get("OPENPYXL_CACHE_DEBUG"):
            import sys
            print(f"[oxlcache] FILL FAILED {type(exc).__name__}: {exc}", file=sys.stderr)
    wb._oxlcache_data_only = data_only
    return wb


def install():
    global _original_load
    import openpyxl
    if getattr(openpyxl, "_oxlcache_installed", False):
        return
    _original_load = openpyxl.load_workbook
    openpyxl.load_workbook = cached_load_workbook
    if _ENABLED and _FASTBIND:
        try:
            from openpyxl.worksheet._reader import WorksheetReader
            compatible = (
                openpyxl.__version__.startswith("3.1")
                and "Cell(self.ws, row=" in inspect.getsource(WorksheetReader.bind_cells)
            )
            if compatible:
                WorksheetReader.bind_cells = _make_fast_bind_cells()
        except Exception:
            pass  # fall back to stock bind_cells (correct, just slower)
    openpyxl._oxlcache_installed = True
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)


install()
