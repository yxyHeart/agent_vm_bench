import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "pylibs"))
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
BASE = os.path.dirname(os.path.abspath(__file__))
BIG = os.path.join(BASE, "template.xlsx")

def small_flow():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 201):
        for c in range(1, 15):
            ws.cell(row=r, column=c, value=(r * c) / 7.0 if c % 3 else "val-%d-%d" % (r, c))
    ws["F1"] = "=SUM(A1:A200)"
    ws.merge_cells("H1:I2")
    ws["H1"].font = Font(bold=True, size=14)
    ws["H1"].fill = PatternFill("solid", start_color="DDDDDD")
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=50))
    ws.add_chart(ch, "K1")
    p = os.path.join(BASE, "corr_small.xlsx")
    wb.save(p)
    wb2 = load_workbook(p)
    wb2["Data"]["A1"] = 42
    wb2.save(p)
    wb3 = load_workbook(p, read_only=True)
    n = sum(1 for row in wb3["Data"].iter_rows() for c in row if c.value is not None)
    wb3.close()
    wb4 = load_workbook(p)
    _ = [c.value for row in wb4["Data"].iter_rows(min_row=1, max_row=60) for c in row]

def big_flow():
    wb = load_workbook(BIG, data_only=False)
    ws = wb["Raw_Sample"]
    for row in ws.iter_rows(min_row=1, max_row=3000):
        for c in row:
            _ = c.value
    wb.save(os.path.join(BASE, "corr_big.xlsx"))

try:
    small_flow()
except Exception:
    pass
try:
    big_flow()
except Exception:
    pass
sys.exit(0)
