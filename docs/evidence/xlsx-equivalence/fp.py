import hashlib
import json
import sys

from openpyxl import load_workbook

P = "/opt/document-bench/xlsx/input/monthly_operations_template.xlsx"


def wb_fingerprint(path, data_only=False):
    wb = load_workbook(path, data_only=data_only)
    info = {"sheets": list(wb.sheetnames)}
    total = hashlib.sha256()
    for sn in wb.sheetnames:
        ws = wb[sn]
        sh = hashlib.sha256()
        n = 0
        for (r, c), cell in ws._cells.items():
            st = cell._style
            st_t = tuple(st) if st is not None else None
            cm = cell._comment
            cm_t = (cm.text[:64], cm.author) if cm is not None else None
            hl = cell._hyperlink.target if getattr(cell, "_hyperlink", None) is not None else None
            sh.update(f"{r}|{c}|{cell.value!r}|{cell.data_type}|{st_t}|{cm_t}|{hl}".encode())
            n += 1
        info[sn] = {
            "cells": n,
            "cellhash": sh.hexdigest()[:16],
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": sorted(str(m) for m in ws.merged_cells.ranges),
            "freeze": str(ws.freeze_panes),
            "charts": len(ws._charts),
            "chart_types": sorted(ch.__class__.__name__ for ch in ws._charts),
            "dv": len(ws.data_validations.dataValidation),
            "cf": len(list(ws.conditional_formatting)),
        }
        total.update(sh.digest())
    info["defined_names"] = sorted(wb.defined_names)
    info["external_links"] = len(wb._external_links)
    info["TOTAL"] = total.hexdigest()
    wb.close()
    return info


if __name__ == "__main__":
    tag = sys.argv[1]
    fp = wb_fingerprint(P)
    fp["tag"] = tag
    print("FPRESULT " + json.dumps(fp, ensure_ascii=False))
